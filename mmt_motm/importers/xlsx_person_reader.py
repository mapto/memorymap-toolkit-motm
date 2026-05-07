import json
import re
from datetime import datetime
from openpyxl import load_workbook


# ==================================================
# Utility functions
# ==================================================

def parse_date(value):
    """
    Convertion '15.6.1921' in 'YYYY-MM-DD'
    """
    try:
        return datetime.strptime(value.strip(), "%d.%m.%Y").date().isoformat()
    except Exception:
        return None


def parse_coordinates(text):
    """
    Extraction of coordinates:
    'Koordinaten: 50.58 / 8.67'
    """
    match = re.search(r"([\d.]+)\s*/\s*([\d.]+)", text)
    if match:
        return {
            "lat": float(match.group(1)),
            "lon": float(match.group(2)),
        }
    return None

def clean_value(value):
    if not value:
        return None

    value = str(value).strip()

    # valori da considerare "vuoti"
    if value in ["", "-", "–", "—"]:
        return None

    return value


# ==================================================
# RAW reader
# ==================================================

def read_raw_person_sheet(xlsx_path, sheet_name=None):
    wb = load_workbook(xlsx_path)

    sheet = wb[sheet_name] if sheet_name else wb.active

    rows = []

    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue

        field = str(row[0]).strip() if len(row) > 0 and row[0] else ""
        value = row[1] if len(row) > 1 else None
        extra = row[2] if len(row) > 2 else None

        rows.append((field, value, extra))

    return rows


# ==================================================
# NORMALIZED reader
# PERSON + birth_place + Interview
# ==================================================

def read_person_record(
    xlsx_path,
    sheet_name="Scheda individuale",
    write_json=False
):
    import os
    import re
    from openpyxl import load_workbook

    rows = read_raw_person_sheet(xlsx_path, sheet_name=sheet_name)

    record = {
        "identifier": None,
        "person": {
            "given_name": None,
            "previous_given_name": None,
            "family_name": None,
            "previous_family_name": None,
            "gender": None,
            "birth_date": None,
            "attributes": {}
        },
        "birth_place": {
            "name": None,
            "coordinates": None,
            "wikidata_id": None,
            "regions": [],
        },
        "interviews": [],
        "family": []
    }

    in_person_section = False
    in_sources_section = False

    # ==================================================
    # MAIN SHEET (PERSON + INTERVIEWS)
    # ==================================================
    for field, value, extra in rows:
        field = field.strip() if field else ""
        
        value_str = clean_value(value)
        extra_str = clean_value(extra)

        # ---------------------------
        # Identifier
        # ---------------------------
        if record["identifier"] is None:
            combined = " ".join([x for x in [field, value_str] if x])
            combined = combined.replace("–", "-")

            match = re.search(r"Metadaten\s*-\s*([A-Za-z0-9_]+)", combined)
            if match:
                record["identifier"] = match.group(1)

        # ---------------------------
        # Section markers
        # ---------------------------
        if field == "PERSON":
            in_person_section = True
            continue

        if field == "BILDUNG UND BERUF":
            in_person_section = False

        if field == "QUELLEN":
            in_sources_section = True
            continue

        if field == "MIGRATION":
            in_sources_section = False

        # ---------------------------
        # PERSON
        # ---------------------------
        if in_person_section:

            if field == "Vorname":
                record["person"]["given_name"] = value_str

            elif field == "ehemaliger Vorname":
                record["person"]["previous_given_name"] = value_str or None

            elif field == "Nachname":
                record["person"]["family_name"] = value_str

            elif field == "ehemaliger Nachname":
                record["person"]["previous_family_name"] = value_str or None

            elif field == "Geschlecht":
                record["person"]["gender"] = value_str

            elif field == "Geburtsdatum":
                record["person"]["birth_date"] = parse_date(value_str)

            elif field == "Verfolgtengruppe(n) NS":
                record["person"]["attributes"]["ns_persecution_group"] = value_str

            elif field == "Geburtsort":
                record["birth_place"]["name"] = value_str

                parts = []

                if extra_str:
                    parts = [p.strip() for p in extra_str.split("|") if p.strip()]
                    
                for part in parts:
                    if part.startswith("Land:"):
                        record["birth_place"]["regions"].append(part.split(":", 1)[1].strip())

                    elif part.startswith("Region:"):
                        record["birth_place"]["regions"].append(part.split(":", 1)[1].strip())

                    elif part.startswith("Kreis:"):
                        record["birth_place"]["regions"].append(part.split(":", 1)[1].strip())

                    elif part.startswith("Koordinaten:"):
                        record["birth_place"]["coordinates"] = parse_coordinates(part)

                    elif part.startswith("Wikidata:"):
                        record["birth_place"]["wikidata_id"] = part.split(":", 1)[1].strip()

        # ---------------------------
        # INTERVIEWS
        # ---------------------------
        if in_sources_section:

            match = re.match(r"Quelle (\d+)\s*–\s*(.+)", field)

            if not match:
                continue

            index = int(match.group(1))
            subfield = match.group(2)

            while len(record["interviews"]) < index:
                record["interviews"].append({
                    "type": None,
                    "place": None,
                    "date": None,
                    "interviewer": None,
                    "archive_id": None,
                })

            current = record["interviews"][index - 1]

            if subfield == "Typ":
                current["type"] = clean_value(value)

            elif subfield == "Ort":
                current["place"] = clean_value(value)

            elif subfield == "Datum":
                current["date"] = parse_date(value_str) if value_str else None

            elif subfield == "Gesprächspartner/in":
                current["interviewer"] = clean_value(value)

            elif subfield == "ID":
                current["archive_id"] = clean_value(value)
    
    record["interviews"] = [
        i for i in record["interviews"]
        if any([
            i["type"],
            i["place"],
            i["date"],
            i["interviewer"],
            i["archive_id"]
        ])
]

    # ==================================================
    # FAMILY SHEET
    # ==================================================
    wb = load_workbook(xlsx_path)

    if "Familie" in wb.sheetnames:
        sheet = wb["Familie"]

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i <= 1:
                continue  # header

            if not row or all(cell is None for cell in row):
                continue

            (
                id_speaker,
                relation,
                given_name,
                family_name,
                birth_date_raw,
                birth_place,
                notes
            ) = row[:7]

            birth_date = None
            if isinstance(birth_date_raw, str) and "–" not in birth_date_raw:
                birth_date = parse_date(birth_date_raw)

            record["family"].append({
                "relation": clean_value(relation),
                "given_name": clean_value(given_name),
                "family_name": clean_value(family_name),
                "birth_date": birth_date,
                "birth_place": clean_value(birth_place),
                "notes": clean_value(notes)
            })

    # ==================================================
    # JSON OUTPUT
    # ==================================================
    if write_json:
        identifier = record.get("identifier", "unknown")

        os.makedirs("data/parsed", exist_ok=True)

        final_path = f"data/parsed/{identifier}.person.json"

        with open(final_path, "w", encoding="utf-8") as f:
            json.dump(record, f, ensure_ascii=False, indent=2)

    return record
